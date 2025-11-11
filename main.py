# app.py
import time
from math import sqrt
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import yfinance as yf
from flask import Flask, jsonify, request, render_template
from pandas.tseries.offsets import BDay
from sklearn.linear_model import Ridge
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from pathlib import Path
LOG_XLSX = Path("predictions_log.xlsx")
app = Flask(__name__)

# ---------------- Config ----------------
LAGS_DEFAULT = 10
BACKTEST_SIZE_DEFAULT = 30
CACHE_TTL_SEC = 15 * 60  # cache price history for 15 minutes

# 50 sample tickers (swap to .NS for NSE if you want)
STOCKS: Dict[str, str] = {
    "Reliance Industries (RELIANCE.NS)": "RELIANCE.NS",
    "TCS (TCS.NS)": "TCS.NS",
    "HDFC Bank (HDFCBANK.NS)": "HDFCBANK.NS",
    "ICICI Bank (ICICIBANK.NS)": "ICICIBANK.NS",
    "Infosys (INFY.NS)": "INFY.NS",
    "ITC (ITC.NS)": "ITC.NS",
    "State Bank of India (SBIN.NS)": "SBIN.NS",
    "Bharti Airtel (BHARTIARTL.NS)": "BHARTIARTL.NS",
    "Hindustan Unilever (HINDUNILVR.NS)": "HINDUNILVR.NS",
    "Larsen & Toubro (LT.NS)": "LT.NS",
    "Axis Bank (AXISBANK.NS)": "AXISBANK.NS",
    "Kotak Mahindra Bank (KOTAKBANK.NS)": "KOTAKBANK.NS",
    "Bajaj Finance (BAJFINANCE.NS)": "BAJFINANCE.NS",
    "Bajaj Finserv (BAJAJFINSV.NS)": "BAJAJFINSV.NS",
    "Adani Enterprises (ADANIENT.NS)": "ADANIENT.NS",
    "Adani Ports (ADANIPORTS.NS)": "ADANIPORTS.NS",
    "Maruti Suzuki (MARUTI.NS)": "MARUTI.NS",
    "Sun Pharma (SUNPHARMA.NS)": "SUNPHARMA.NS",
    "Asian Paints (ASIANPAINT.NS)": "ASIANPAINT.NS",
    "Mahindra & Mahindra (M&M.NS)": "M&M.NS",
    "Tata Motors (TATAMOTORS.NS)": "TATAMOTORS.NS",
    "NTPC (NTPC.NS)": "NTPC.NS",
    "Power Grid (POWERGRID.NS)": "POWERGRID.NS",
    "UltraTech Cement (ULTRACEMCO.NS)": "ULTRACEMCO.NS",
    "JSW Steel (JSWSTEEL.NS)": "JSWSTEEL.NS",
    "Tata Steel (TATASTEEL.NS)": "TATASTEEL.NS",
    "Nestlé India (NESTLEIND.NS)": "NESTLEIND.NS",
    "Titan (TITAN.NS)": "TITAN.NS",
    "Tech Mahindra (TECHM.NS)": "TECHM.NS",
    "Wipro (WIPRO.NS)": "WIPRO.NS",
    "Grasim (GRASIM.NS)": "GRASIM.NS",
    "Cipla (CIPLA.NS)": "CIPLA.NS",
    "Divi's Laboratories (DIVISLAB.NS)": "DIVISLAB.NS",
    "Dr. Reddy's (DRREDDY.NS)": "DRREDDY.NS",
    "Britannia (BRITANNIA.NS)": "BRITANNIA.NS",
    "Eicher Motors (EICHERMOT.NS)": "EICHERMOT.NS",
    "Hero MotoCorp (HEROMOTOCO.NS)": "HEROMOTOCO.NS",
    "Bajaj Auto (BAJAJ-AUTO.NS)": "BAJAJ-AUTO.NS",
    "Tata Consumer (TATACONSUM.NS)": "TATACONSUM.NS",
    "HDFC Life (HDFCLIFE.NS)": "HDFCLIFE.NS",
    "SBI Life (SBILIFE.NS)": "SBILIFE.NS",
    "ICICI Lombard (ICICIGI.NS)": "ICICIGI.NS",
    "Hindalco (HINDALCO.NS)": "HINDALCO.NS",
    "Coal India (COALINDIA.NS)": "COALINDIA.NS",
    "ONGC (ONGC.NS)": "ONGC.NS",
    "LTIMindtree (LTIM.NS)": "LTIM.NS",
    "Apollo Hospitals (APOLLOHOSP.NS)": "APOLLOHOSP.NS",
    "Avenue Supermarts (DMART.NS)": "DMART.NS",
    "DLF (DLF.NS)": "DLF.NS",
    "Godrej Consumer (GODREJCP.NS)": "GODREJCP.NS",
    "BSE Limited (BSE.NS)": "BSE.NS",
    "UPL (UPL.NS)": "UPL.NS",
    "Dixon Technologies (DIXON.NS)": "DIXON.NS",
    "Suzlon Energy (SUZLON.NS)": "SUZLON.NS",
    "Coforge (COFORGE.NS)": "COFORGE.NS",
    "Brigade Enterprises (BRIGADE.NS)": "BRIGADE.NS",
    "JSW Steel (JSWSTEEL.NS)": "JSWSTEEL.NS",
}

# ---------------- Price cache ----------------
_price_cache: Dict[str, Tuple[float, pd.DataFrame]] = {}

def _now() -> float:
    return time.time()

def get_cached_history(ticker: str, period: str = "5y") -> pd.DataFrame:
    cached = _price_cache.get(ticker)
    if cached and (_now() - cached[0] < CACHE_TTL_SEC):
        return cached[1].copy()
    df = yf.download(ticker, period=period, interval="1d", auto_adjust=True, progress=False)
    if df is None or df.empty:
        raise ValueError(f"No price data for {ticker}")
    df = df.dropna()
    _price_cache[ticker] = (_now(), df.copy())
    return df

def forecast_multi_step(close: pd.Series, lags: int, steps: int = 3) -> list[float]:
    X, y, _ = make_supervised(close, lags)
    mdl = base_model().fit(X, y)
    hist = close.to_numpy().astype(float).tolist()
    out = []
    for _ in range(steps):
        x = np.array(hist[-lags:])[::-1].reshape(1, -1)
        yhat = float(mdl.predict(x)[0])
        out.append(yhat)
        hist.append(yhat)
    return out

# ---------------- Feature helpers ----------------
def get_close_series(df: pd.DataFrame) -> pd.Series:
    """Return Close as 1-D float Series (avoids (N,1) and FutureWarnings)."""
    s = df["Close"] if "Close" in df.columns else df.iloc[:, 0]
    if isinstance(s, pd.DataFrame):
        s = s.squeeze("columns")
    arr = pd.to_numeric(s.to_numpy().reshape(-1), errors="coerce")
    ser = pd.Series(arr, index=df.index[-len(arr):], dtype="float64").dropna()
    if ser.empty:
        raise ValueError("Close series is empty after cleaning.")
    return ser

def base_model() -> Pipeline:
    return Pipeline([("scaler", StandardScaler()), ("reg", Ridge(alpha=1.0))])

def make_supervised(close: pd.Series, lags: int) -> tuple[np.ndarray, np.ndarray, pd.DatetimeIndex]:
    close = pd.Series(close, dtype="float64")
    df = pd.DataFrame({"y": close})
    for i in range(1, lags + 1):
        df[f"lag_{i}"] = df["y"].shift(i)
    df = df.dropna()
    if df.empty:
        raise ValueError("Not enough data after creating lags.")
    X = df[[f"lag_{i}" for i in range(1, lags + 1)]].to_numpy()  # (n, lags)
    y = df["y"].to_numpy()                                       # (n,)
    return X, y, df.index

def fit_and_predict_next(close: pd.Series, lags: int) -> float:
    if len(close) <= lags + 20:
        raise ValueError("Not enough history (need > lags+20 days).")
    X, y, _ = make_supervised(close, lags)
    mdl = base_model().fit(X, y)
    last_lags = close.iloc[-1:-(lags + 1):-1].to_numpy()  # shape (lags,)
    if last_lags.ndim != 1 or last_lags.shape[0] < lags:
        raise ValueError(f"Bad last_lags shape {last_lags.shape}; expected ({lags},)")
    return float(mdl.predict(last_lags.reshape(1, -1))[0])

def backtest_nextday(close: pd.Series, lags: int, test_size: int) -> pd.DataFrame:
    X, y, idx = make_supervised(close, lags)
    n = len(y)
    if n < 30:
        raise ValueError("Not enough samples for backtest (need >= 30).")
    test_size = max(5, min(test_size, n // 2))
    start = n - test_size
    rows: List[Dict] = []
    for t in range(start, n):
        X_train, y_train = X[:t], y[:t]
        x_test = X[t].reshape(1, -1)
        prev_close = float(X[t, 0])  # lag_1 == y_{t-1}
        pred = float(base_model().fit(X_train, y_train).predict(x_test)[0])
        actual = float(y[t])
        rows.append({
            "date": idx[t].date(),
            "pred_close": pred,
            "actual_close": actual,
            "error": actual - pred,
            "prev_close": prev_close,
            "direction_ok": np.sign(pred - prev_close) == np.sign(actual - prev_close),
        })
    return pd.DataFrame(rows)

def metrics_from_backtest(bt: pd.DataFrame) -> Dict[str, float]:
    err = bt["error"].to_numpy()
    rmse = sqrt(float(np.mean(err**2)))
    mae = float(np.mean(np.abs(err)))
    mape = float(np.mean(np.abs(err) / np.maximum(1e-9, bt["actual_close"].to_numpy()))) * 100.0
    hit = float(bt["direction_ok"].mean()) * 100.0
    return {"rmse": rmse, "mae": mae, "mape_pct": mape, "hit_rate_pct": hit}

def interval_from_residuals(pred: float, residuals: pd.Series,
                            q_low: float = 0.10, q_high: float = 0.90) -> tuple[float, float, List[float]]:
    residuals = pd.Series(residuals, dtype="float64").dropna()
    if len(residuals) >= 15:
        lo = pred + float(residuals.quantile(q_low))
        hi = pred + float(residuals.quantile(q_high))
    else:
        s = float(residuals.std(ddof=1)) if len(residuals) > 1 else 0.0
        lo, hi = pred - 1.64 * s, pred + 1.64 * s
    if lo > hi:
        lo, hi = hi, lo
    band10 = list(np.linspace(lo, hi, 10))
    return float(lo), float(hi), [float(x) for x in band10]

# ---------------- Routes ----------------
@app.get("/")
def ui():
    # serves templates/ui.html (place your HTML there)
    return render_template("ui.html")

@app.get("/api/stocks")
def api_stocks():
    return jsonify({"stocks": [{"label": k, "ticker": v} for k, v in STOCKS.items()]})

@app.get("/api/predict")
def api_predict():
    raw = request.args.get("ticker", type=str)
    if not raw:
        return jsonify({"error": "ticker is required"}), 400

    symbol = STOCKS.get(raw, raw)  # allow label or raw symbol
    lags = request.args.get("lags", default=LAGS_DEFAULT, type=int)
    backtest_size = request.args.get("backtest_size", default=BACKTEST_SIZE_DEFAULT, type=int)
    q_low = request.args.get("q_low", default=0.10, type=float)
    q_high = request.args.get("q_high", default=0.90, type=float)

    try:
        # --- compute prediction & artifacts (unchanged) ---
        df = get_cached_history(symbol, period="5y")
        close = get_close_series(df)

        last_close = float(close.iloc[-1])
        last_date = df.index[-1].date()
        pred_date = (df.index[-1] + BDay(1)).date()  # weekend-aware

        pred_value = fit_and_predict_next(close, lags=lags)
        bt = backtest_nextday(close, lags=lags, test_size=backtest_size).sort_values("date", ascending=False)
        metrics = metrics_from_backtest(bt)
        lo, hi, band10 = interval_from_residuals(pred_value, bt["error"], q_low, q_high)

        rows = [{
            "date": r["date"].isoformat(),
            "pred_close": round(float(r["pred_close"]), 2),
            "actual_close": round(float(r["actual_close"]), 2),
            "error": round(float(r["error"]), 2),
            "direction_ok": bool(r["direction_ok"]),
        } for _, r in bt.iterrows()]

        # --- build response first, then log to Excel ---
        payload = {
            "ticker": symbol,
            "last_close": round(last_close, 2),
            "last_date": last_date.isoformat(),
            "pred_date": pred_date.isoformat(),
            "prediction": round(float(pred_value), 2),
            "metrics": {
                "rmse": round(metrics["rmse"], 4),
                "mae": round(metrics["mae"], 4),
                "mape_pct": round(metrics["mape_pct"], 2),
                "hit_rate_pct": round(metrics["hit_rate_pct"], 1),
            },
            "interval": {
                "lo": round(lo, 2),
                "hi": round(hi, 2),
                "band10": [round(x, 2) for x in band10],
            },
            "backtest_rows": rows,
            # include inputs used so they’re logged too
            "lags": lags,
            "backtest_size": backtest_size,
        }

        # --- append a single row to Excel ---
        # requires: from pathlib import Path; LOG_XLSX = Path("predictions_log.xlsx")
        row_df = pd.DataFrame([{
            "ts": pd.Timestamp.now(tz="Asia/Kolkata").strftime("%Y-%m-%d %H:%M:%S"),
            "ticker": payload["ticker"],
            "last_date": payload["last_date"],
            "pred_date": payload["pred_date"],
            "last_close": payload["last_close"],
            "prediction": payload["prediction"],
            "lo": payload["interval"]["lo"],
            "hi": payload["interval"]["hi"],
            "rmse": payload["metrics"]["rmse"],
            "mape_pct": payload["metrics"]["mape_pct"],
            "hit_rate_pct": payload["metrics"]["hit_rate_pct"],
            "lags": payload["lags"],
            "backtest_size": payload["backtest_size"],
        }])

        if LOG_XLSX.exists():
            # ensure openpyxl engine is available
            from openpyxl import load_workbook  # noqa: F401
            with pd.ExcelWriter(LOG_XLSX, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = "Sheet1"
                if sheet in writer.book.sheetnames:
                    startrow = writer.book[sheet].max_row
                else:
                    writer.book.create_sheet(sheet)
                    startrow = 0
                row_df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=startrow)
        else:
            with pd.ExcelWriter(LOG_XLSX, engine="openpyxl") as writer:
                row_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # --- return as usual ---
        return jsonify(payload)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/predict_1h")
def api_predict_1h():
    symbol = request.args.get("ticker", type=str)
    lags = request.args.get("lags", default=LAGS_DEFAULT, type=int)
    steps = request.args.get("steps", default=3, type=int)

    df = get_history_for_rsi(symbol, interval="1h", period="1y")
    close = get_close_series(df)

    preds = forecast_multi_step(close, lags=lags, steps=steps)
    last_ts = df.index[-1]
    horizon = [(last_ts + pd.Timedelta(hours=i+1)).isoformat() for i in range(steps)]

    return jsonify({
        "ticker": symbol,
        "interval": "1h",
        "steps": steps,
        "last_ts": last_ts.isoformat(),
        "horizon": horizon,
        "predictions": [round(p, 2) for p in preds]
    })


# --- ADD: RSI helper ---


_rsi_cache = {}

def get_history_for_rsi(ticker: str, interval: str = "1h", period: str = "1y") -> pd.DataFrame:
    key = (ticker, interval, period)
    cached = _rsi_cache.get(key)
    if cached and (time.time() - cached[0] < 60):   # 60s TTL
        return cached[1].copy()

    df = yf.download(ticker, period=period, interval=interval, auto_adjust=True, progress=False)
    if df is None or df.empty:
        raise ValueError(f"No intraday data for {ticker}")
    df = df.dropna()
    _rsi_cache[key] = (time.time(), df.copy())
    return df


def compute_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    """
    Wilder-style RSI. Returns a Series aligned with `close`.
    """
    close = pd.Series(close, dtype="float64")
    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = -delta.clip(upper=0.0)

    # Wilder's smoothing (RMA)
    avg_gain = gain.ewm(alpha=1/period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, adjust=False).mean().replace(0, np.nan)

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.fillna(50.0)


# --- ADD: Live RSI API ---
@app.get("/api/rsi")
def api_rsi():
    """
    Returns RSI for all STOCKS, sorted high -> low.
    Flags RSI < 35 or > 75 for red highlighting on the UI.
    Query params:
      - period: int (default 14)
    """
    period = request.args.get("period", default=14, type=int)
    low_thr, high_thr = 30.0, 70.0

    # inside: @app.get("/api/rsi")
    rows = []
    period = request.args.get("period", default=14, type=int)
    low_thr, high_thr = 30.0, 70.0

    for label, symbol in STOCKS.items():
        try:
            # use a short window to speed up first load and reduce rate-limit risk
            df = get_history_for_rsi(symbol, interval="1h", period="1y")
            close = get_close_series(df)
            rsi_val = float(compute_rsi(close, period=period).iloc[-1])

            rows.append({
                "label": label,
                "ticker": symbol,
                "rsi": round(rsi_val, 2),
                "last_close": round(float(close.iloc[-1]), 2),
                "last_date": df.index[-1].date().isoformat(),
                "alert": (rsi_val < low_thr) or (rsi_val > high_thr)
            })
        except Exception:
            # skip bad/temporarily failing tickers instead of aborting the whole response
            continue

    rows.sort(key=lambda x: x["rsi"], reverse=True)
    return jsonify({
        "period": period,
        "low_threshold": low_thr,
        "high_threshold": high_thr,
        "count": len(rows),
        "rows": rows,
        "ts": int(time.time())
    })


# --- ADD: RSI UI route ---
@app.get("/rsi")
def rsi_page():
    return render_template("rsi.html")


# ---------------- Main ----------------
if __name__ == "__main__":
    app.run(debug=True)
